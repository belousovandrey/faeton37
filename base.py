import openpyxl

# d = input('Ввести путь для донора:')
d = 'donor.xlsx'
# f = input('Ввести путь для финального файла:')
f = 'final.xlsx'
i = 2
wb = openpyxl.load_workbook(d)
sh = wb.active
wb_out = openpyxl.Workbook()
shout = wb_out.active
for row in range(sh.max_row - 2, -1, -3):  # для фэтона
    a = sh[row][0].value
    b = str(sh[row + 1][2].value)[:10]
    parts = str(sh[row][3].value) + ',' + str(sh[row + 1][3].value) + str(sh[row + 2][3].value)
    price = sh[row][18].value
    shout.cell(row=i, column=1).value = a
    shout.cell(row=i, column=2).value = b
    shout.cell(row=i, column=3).value = parts
    shout.cell(row=i, column=4).value = price
    i += 1
    wb_out.save(f)
    print(i)
wb.close()
wb_out.close()
